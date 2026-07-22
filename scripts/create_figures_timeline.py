from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Users\TS6207_VENKAT\Projects\Post-Meeting Summary Bot\outputs\dates_filled\dates_filled_blue.xlsx")
OUTPUT = Path(r"C:\Users\TS6207_VENKAT\Projects\Post-Meeting Summary Bot\outputs\dates_filled\dates_timeline.xlsx")


ACHIEVEMENTS = {
    "Tutankhamun": "Egyptian pharaoh whose intact tomb transformed modern knowledge of New Kingdom burial culture.",
    "Ozymandias": "Ramesses II, remembered for monumental building projects and long Egyptian rule.",
    "Taigong Wang": "Strategist traditionally credited with helping found the Zhou dynasty.",
    "Solomon": "Biblical king associated with wisdom, the First Temple, and Israel's golden age.",
    "Romulus": "Legendary founder and first king of Rome.",
    "Leonidas": "Led the Spartan stand at Thermopylae.",
    "Darius III": "Last Achaemenid king, defeated by Alexander the Great.",
    "Ptolemaios": "Founded the Ptolemaic dynasty in Egypt after Alexander's conquests.",
    "Alexander": "Created one of antiquity's largest empires and spread Hellenistic culture.",
    "Ashoka": "Mauryan emperor known for Buddhist patronage and edict inscriptions.",
    "Shi Huang Di": "Unified China and became its first emperor.",
    "Xu Fu": "Legendary Qin-era seeker of immortality islands.",
    "Consort Yu": "Remembered for loyalty to Xiang Yu in the Chu-Han contention.",
    "Xiang Yu": "Military leader who challenged Liu Bang after the Qin collapse.",
    "Spartacus": "Led a major slave uprising against Rome.",
    "Julius Caesar": "Roman general and dictator whose reforms and assassination reshaped Rome.",
    "Cleopatra": "Last active Ptolemaic ruler of Egypt.",
    "Caligula": "Roman emperor remembered for controversial rule.",
    "Boudica": "Led the Iceni revolt against Roman Britain.",
    "Nero Claudius": "Roman emperor associated with the Great Fire of Rome and Julio-Claudian collapse.",
    "Yamato Takeru": "Legendary Japanese prince and warrior.",
    "Lu Bu Fengxian": "Three Kingdoms-era warrior famed for martial prowess.",
    "Himiko": "Shaman-queen of Yamatai described in Chinese records.",
    "Sima Yi": "Strategist whose family paved the way for the Jin dynasty.",
    "Zhuge Liang": "Shu Han chancellor and strategist idealized for statecraft.",
    "Iyo": "Successor queen of Yamatai after Himiko in early Japanese records.",
    "Zenobia": "Queen of Palmyra who challenged Roman authority.",
    "Georgios": "Saint George, martyr and later dragon-slayer legend.",
    "Atilla": "King of the Huns who threatened both Roman empires.",
    "Kriemhild": "Central heroine of the Nibelungenlied revenge cycle.",
    "King Arthur": "Legendary British king at the center of the Matter of Britain.",
    "Lanling Wang": "Northern Qi prince famed for battlefield bravery.",
    "Xuanzang Sanzang": "Buddhist monk whose pilgrimage to India shaped East Asian Buddhism.",
    "Wu Zetian": "Only woman to rule China as emperor in her own name.",
    "Charlemagne": "Carolingian ruler crowned emperor and associated with European consolidation.",
    "Roland": "Frankish hero of the Battle of Roncevaux Pass.",
    "Ono no Komachi": "Celebrated Heian poet and beauty.",
    "Minamoto no Raiko": "Legendary warrior credited with monster-slaying tales.",
    "Watanabe no Tsuna": "Raiko retainer famous in oni-slaying legends.",
    "Sei Shonagon": "Author of The Pillow Book.",
    "Murasaki Shikibu": "Author of The Tale of Genji.",
    "Avicebron": "Jewish philosopher and poet known for Neoplatonic thought.",
    "Old Man of the Mountain": "Hasan-i Sabbah, founder of the Nizari Ismaili state at Alamut.",
    "Minamoto no Tametomo": "Heian warrior known for legendary archery.",
    "Musashibo Benkei": "Warrior monk famed for loyalty to Minamoto no Yoshitsune.",
    "Richard the Lionheart": "Crusader king of England and commander in the Third Crusade.",
    "Ushiwakamaru": "Youth name of Minamoto no Yoshitsune, famed Genpei War commander.",
    "John Lackland": "King John, associated with Magna Carta.",
    "Jacques de Molay": "Last grand master of the Knights Templar.",
    "William Tell": "Swiss legendary archer and resistance figure.",
    "Constantine XI": "Last Byzantine emperor, killed at Constantinople's fall.",
    "Gilles de Rais": "French marshal and companion of Jeanne d'Arc, later executed for crimes.",
    "Jeanne d’Arc": "Led French forces in the Hundred Years' War and became a national saint.",
    "Francois Prelati": "Alchemist linked to the Gilles de Rais trials.",
    "Vlad": "Vlad III Dracula, Wallachian ruler known for anti-Ottoman campaigns.",
    "Christopher Colombus": "Navigator whose 1492 voyage linked Europe and the Americas.",
    "Leonardo da Vinci": "Renaissance artist, engineer, and scientist.",
    "Paracelsus von Hohenheim": "Physician who helped reshape early modern medicine and alchemy.",
    "Kato Danzo": "Legendary ninja and illusionist.",
    "Taketa Shingen": "Sengoku daimyo known for military reforms and rivalry with Uesugi Kenshin.",
    "Sen no Rikyo": "Tea master who codified wabi-cha aesthetics.",
    "Kashin Koji": "Legendary illusionist and magician of Sengoku tales.",
    "Nagao Kagetora": "Uesugi Kenshin, daimyo famed as the Dragon of Echigo.",
    "Ivan the Terrible": "First tsar of Russia.",
    "Oda Nobunaga": "Daimyo who began Japan's late-Sengoku unification.",
    "Saika Magoichi": "Name associated with leaders of the Saika mercenary gunmen.",
    "Francis Drake": "English privateer and circumnavigator.",
    "Sugitani Zenjuboto": "Marksman remembered for attempting to assassinate Oda Nobunaga.",
    "Don Quixote": "Literary knight whose novel became a landmark of modern fiction.",
    "Oda Nobukatsu": "Sengoku lord and son of Oda Nobunaga.",
    "Elizabeth Bathory": "Hungarian noblewoman infamous for murder accusations.",
    "Fuma Kotaro": "Legendary leader of the Fuma ninja clan.",
    "William Shakespeare": "Playwright and poet central to English literature.",
    "Mysterious Ranmaru X": "Fictionalized version of Mori Ranmaru in Fate-related material.",
    "Chacha": "Yodo-dono, political figure in late Sengoku and Osaka Castle conflicts.",
    "Yagyu Munenori": "Swordmaster and Tokugawa military instructor.",
    "Qin Llangyu": "Ming loyalist general known for battlefield command.",
    "Izumo no Okuni": "Performer traditionally credited with founding kabuki.",
    "Miyamoto Musashi": "Swordsman and author of The Book of Five Rings.",
    "Sasaki Kojiro": "Swordsman famous for his duel with Miyamoto Musashi.",
    "Hozoin Inshun": "Spear master of the Hozoin school.",
    "Yui Shosetsu": "Military strategist linked to the Keian uprising.",
    "Miyamoto Iori": "Adopted son and heir of Miyamoto Musashi.",
    "Amakusa Shiro": "Led the Shimabara Rebellion.",
    "Edward Teach": "Blackbeard, notorious Atlantic pirate.",
    "Mary Read": "Pirate associated with Anne Bonny and Calico Jack.",
    "Abigail Williams": "Accuser in the Salem witch trials.",
    "Batholomew Roberts": "Bartholomew Roberts, one of the most successful Golden Age pirates.",
    "Anne Bonny": "Pirate active in the Caribbean alongside Mary Read.",
    "Chevalier d’Eon": "Diplomat, spy, soldier, and noted gender-nonconforming public figure.",
    "Charles-Henri Sanson": "Executioner of Paris during the French Revolution.",
    "Antonio Salieri": "Composer and imperial Kapellmeister in Vienna.",
    "Marie Antoinette": "Queen of France executed during the Revolution.",
    "Mozart": "Composer whose works define the Classical era.",
    "Katsuhika Hokusai": "Artist of The Great Wave and major ukiyo-e innovator.",
    "Kyokutei Bakin": "Author of Nanso Satomi Hakkenden.",
    "Charlotte Corday": "Assassinated Jean-Paul Marat.",
    "Napoleon": "French emperor and military reformer whose campaigns transformed Europe.",
    "Charles Babbage": "Designed the Difference Engine and Analytical Engine.",
    "Edmond Dantes": "Fictional protagonist of The Count of Monte Cristo, noted for revenge and reinvention.",
    "Mary Anning": "Pioneering fossil collector and paleontological discoverer.",
    "Hans Christian Andersen": "Author of influential literary fairy tales.",
    "Florence Nightingale": "Founded modern nursing and hospital statistics reform.",
    "Geronimo": "Apache leader who resisted U.S. and Mexican expansion.",
    "Helena Blavatsky": "Co-founder of the Theosophical Society.",
    "Yamanami Keisuke": "Shinsengumi officer.",
    "Kondo Isami": "Commander of the Shinsengumi.",
    "Kawakami Gensai": "Bakumatsu assassin counted among the Four Hitokiri.",
    "Hijikata Toshizo": "Vice commander of the Shinsengumi.",
    "Sakamoto Ryoma": "Brokered alliances that helped bring about the Meiji Restoration.",
    "Okada Izo": "Bakumatsu assassin associated with Tosa loyalists.",
    "Takasugi Shinsaku": "Founded the Kiheitai militia.",
    "Nagakura Shinpachi": "Shinsengumi captain and later memoirist.",
    "Harada Sanosuke": "Shinsengumi unit captain.",
    "Saito Hajime": "Shinsengumi captain who survived into the Meiji era.",
    "Todo Heisuke": "Shinsengumi officer.",
    "Okita Soji": "Shinsengumi first unit captain.",
    "Thomas Edison": "Inventor and industrial researcher tied to electric light and phonograph development.",
    "Calamity Jane": "American frontier scout and performer.",
    "Van Gogh": "Post-Impressionist painter whose work shaped modern art.",
    "Nikola Tesla": "Inventor associated with AC power systems.",
    "Billy the Kid": "American outlaw of the Lincoln County War.",
    "Li Shuwen": "Bajiquan martial artist famed for spear and empty-hand skill.",
    "Grigori Rasputin": "Mystic influential at the late Romanov court.",
    "Mata Hari": "Dancer executed as a spy during World War I.",
    "Jack the Ripper (the murders)": "Unidentified killer associated with the 1888 Whitechapel murders.",
    "Anastasia": "Grand Duchess Anastasia Nikolaevna, Romanov daughter killed in 1918.",
    "Voyager": "NASA spacecraft launched in 1977 and now in interstellar mission operations.",
    "Dante Aligheri": "Dante Alighieri, author of the Divine Comedy.",
    "Alessandro di Cagliostro": "Occultist and adventurer linked to European esotericism and scandal.",
    "Yang Gufei": "Yang Guifei, Tang imperial consort central to An Lushan-era memory.",
    "Uesugi Kenshin": "Sengoku daimyo famed for campaigns against Takeda Shingen.",
    "Tenochtitlan": "Aztec capital founded in 1325 and conquered by Spain in 1521.",
    "Sherlock Holmes": "Fictional detective whose methods shaped detective fiction.",
    "Mori Nagayoshi": "Sengoku warrior and Oda/Toyotomi retainer.",
    "Tomoe Gozen": "Onna-musha remembered in Genpei War literature.",
    "Trung Sisters": "Led a Vietnamese revolt against Han rule.",
    "Mochizuki Chiyome": "Legendary woman ninja and spy organizer.",
    "Zhang Jue": "Led the Yellow Turban Rebellion.",
    "Chen Gong": "Strategist associated with Lu Bu.",
    "Jing Ke": "Attempted to assassinate the Qin king who became Qin Shi Huang.",
    "Eric Bloodaxe": "Norse king associated with Norway and Northumbria.",
    "Martha": "Biblical saint associated with hospitality and witness traditions.",
    "Sakata Kintoki": "Folkloric strongman Kintaro, retainer of Minamoto no Raiko.",
    "Tawara Tota": "Fujiwara no Hidesato, hero of giant centipede legend.",
}


TRADITION_HINTS = [
    ("Hassan", "Assassin/Ismaili-inspired legendary material"),
    ("Mysterious", "Fate-series fictional material"),
    ("Einzbern", "Fate-series fictional material"),
    ("Edelfelt", "Fate-series fictional material"),
    ("Aozaki", "Type-Moon fictional material"),
    ("Kuonji", "Type-Moon fictional material"),
    ("Kishinami", "Fate-series fictional material"),
    ("BB", "Fate/Extra fictional material"),
    ("Meltryllis", "Fate/Extra fictional material"),
    ("Kingprotea", "Fate/Extra fictional material"),
    ("Kazuradrop", "Fate/Extra fictional material"),
    ("Passionlip", "Fate/Extra fictional material"),
    ("Sesshoin", "Fate-series fictional material"),
    ("Archetype", "Tsukihime/Type-Moon fictional material"),
]


def achievement_for(name):
    if name in ACHIEVEMENTS:
        return ACHIEVEMENTS[name]
    for token, tradition in TRADITION_HINTS:
        if token in name:
            return f"Character associated with {tradition}; included as a fictional/modern media timeline entry."
    return "Mythological, legendary, literary, or fictional figure; included for cultural influence and associated narrative tradition."


def number_or_none(value):
    return value if isinstance(value, (int, float)) else None


def display_year(year):
    if year is None:
        return "Undated"
    year = int(year)
    return f"{abs(year)} BCE" if year < 0 else f"{year} CE"


def event_year(birth, death):
    b = number_or_none(birth)
    d = number_or_none(death)
    if b is not None and d is not None:
        return round((b + d) / 2)
    return b if b is not None else d


def note_for(birth, death):
    if birth == "-" and death == "-":
        return "No defensible dated lifespan; retained as undated."
    if death == "-":
        return "No confirmed death/deceased date, or figure is immortal/fictional/ongoing."
    if birth == "-":
        return "Birth year unknown; timeline uses death/event year."
    if isinstance(birth, (int, float)) and isinstance(death, (int, float)):
        return "Placed at midpoint of recorded lifespan/active tradition for sorting."
    return "Approximate or tradition-based date."


def main():
    wb = load_workbook(BASE)
    source = wb.active
    if "Timeline" in wb.sheetnames:
        del wb["Timeline"]
    ws = wb.create_sheet("Timeline")

    headers = ["Sort Year", "Display Year", "Figure", "Notable achievement / event", "Birth Year", "Death Year", "Certainty / notes"]
    ws.append(headers)

    rows = []
    for r in range(1, source.max_row + 1):
        name = source.cell(r, 1).value
        if not name:
            continue
        birth = source.cell(r, 2).value
        death = source.cell(r, 3).value
        year = event_year(birth, death)
        rows.append((year, display_year(year), name, achievement_for(str(name)), birth, death, note_for(birth, death)))

    rows.sort(key=lambda row: (row[0] is None, 10**9 if row[0] is None else row[0], row[2]))
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [12, 14, 28, 88, 12, 12, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].value is None:
            row[0].value = ""

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    # Keep the raw sort helper available but visually de-emphasized.
    for cell in ws["A"]:
        cell.font = Font(color="666666", bold=(cell.row == 1))

    wb.save(OUTPUT)
    print(f"saved={OUTPUT}")
    print(f"timeline_rows={len(rows)}")


if __name__ == "__main__":
    main()
