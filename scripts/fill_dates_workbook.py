from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


INPUT = Path(r"C:\Users\TS6207_VENKAT\Downloads\dates.xlsx")
OUTPUT_DIR = Path(r"C:\Users\TS6207_VENKAT\Projects\Post-Meeting Summary Bot\outputs\dates_filled")
OUTPUT = OUTPUT_DIR / "dates_filled_blue.xlsx"


# Only cells that are blank in the source workbook are filled.
# Values here are intentionally conservative: well-attested historical dates
# where available, approximate single-year anchors for legendary settings where
# useful for sorting, and "-" where no defensible dated lifespan exists or the
# entity is non-mortal/not canonically deceased.
KNOWN = {
    "Iyo": {"death": 287},
    "King Arthur": {"death": 537},
    "Kashin Koji": {"death": "-"},
    "Siegfried": {"birth": -500, "death": -450},
    "Fergus mac Roach": {"birth": -50, "death": 30},
    "Mordred": {"birth": 500, "death": 537},
    "Ryogi Shiki": {"birth": 1980, "death": "-"},
    "Rama": {"birth": -5114, "death": -5075},
    "Lancelot": {"birth": 500, "death": 542},
    "Gawain": {"birth": 500, "death": 537},
    "Bedivere": {"birth": 500, "death": "-"},
    "Queen Medb": {"birth": -50, "death": 30},
    "Diarmuid Ua Duihbne": {"birth": 150, "death": 190},
    "Beni-Enma": {"birth": "-", "death": "-"},
    "Suzuka Gozen": {"birth": 900, "death": "-"},
    "Frankenstein": {"birth": 1818, "death": "-"},
    "Sigurd": {"birth": -500, "death": -450},
    "Jason": {"birth": -1300, "death": -1250},
    "Astolfo": {"birth": 750, "death": "-"},
    "Dioscuri": {"birth": -1250, "death": "-"},
    "Tomoe Gozen": {"birth": 1157, "death": 1247},
    "Ibuki Doji": {"birth": "-", "death": "-"},
    "Karna": {"birth": -3138, "death": -3067},
    "Senji Muramasa": {"birth": 1501, "death": 1573},
    "Barghest": {"birth": "-", "death": "-"},
    "Trung Sisters": {"birth": 12, "death": 43},
    "Gareth": {"birth": 500, "death": 537},
    "Cu Chullain": {"birth": -30, "death": 1},
    "Medusa": {"birth": -1250, "death": -1220},
    "Theseus": {"birth": -1250, "death": -1200},
    "Kurohime": {"birth": 1500, "death": "-"},
    "Passionlip": {"birth": 2030, "death": "-"},
    "Lord Logress": {"birth": 450, "death": "-"},
    "Emiya": {"birth": 1980, "death": "-"},
    "David": {"birth": -1040, "death": -970},
    "Gilgamesh": {"birth": -2800, "death": -2700},
    "Robin Hood": {"birth": 1160, "death": 1247},
    "Atlanta": {"birth": -1250, "death": "-"},
    "Euryale": {"birth": "-", "death": "-"},
    "Arash": {"birth": 530, "death": 570},
    "Orion": {"birth": -1250, "death": "-"},
    "Arjuna": {"birth": -3138, "death": -3067},
    "Tristan": {"birth": 530, "death": 570},
    "Tawara Tota": {"birth": 891, "death": 958},
    "Chloe von Einzbern": {"birth": 1999, "death": "-"},
    "Ishtar": {"birth": "-", "death": "-"},
    "James Moriarty": {"birth": 1846, "death": 1891},
    "Asagami Fujino": {"birth": 1980, "death": "-"},
    "Chiron": {"birth": -1250, "death": "-"},
    "Aswatthaman": {"birth": -3138, "death": "-"},
    "Paris": {"birth": -1250, "death": -1200},
    "Osakabehime": {"birth": 1500, "death": "-"},
    "Artemis": {"birth": "-", "death": "-"},
    "Ilyasviel von Einzbern": {"birth": 1986, "death": "-"},
    "Baobhan Sith": {"birth": "-", "death": "-"},
    "Durga": {"birth": "-", "death": "-"},
    "Tiamat": {"birth": "-", "death": "-"},
    "Hector": {"birth": -1250, "death": -1184},
    "Scathach": {"birth": -50, "death": "-"},
    "Fionn mac Cumhaill": {"birth": 180, "death": 283},
    "Brynhildr": {"birth": -500, "death": -450},
    "Tamamo no Mae": {"birth": 1100, "death": 1153},
    "Kiyohime": {"birth": 900, "death": 930},
    "Enkidu": {"birth": -2800, "death": -2700},
    "Jaguar Man": {"birth": "-", "death": "-"},
    "Parvati": {"birth": "-", "death": "-"},
    "Nezha": {"birth": -1100, "death": "-"},
    "Ereshkigal": {"birth": "-", "death": "-"},
    "Ibaraki Doji": {"birth": 950, "death": "-"},
    "Bradamante": {"birth": 760, "death": "-"},
    "Mysterious Alter Ego A": {"birth": "-", "death": "-"},
    "Caenis": {"birth": -1250, "death": "-"},
    "Utsumi Erice": {"birth": 2010, "death": "-"},
    "Vritra": {"birth": "-", "death": "-"},
    "Melusine": {"birth": "-", "death": "-"},
    "Percival": {"birth": 500, "death": "-"},
    "Britomart": {"birth": 1580, "death": "-"},
    "Bhima": {"birth": -3138, "death": -3067},
    "Dorbrynya Nikitich": {"birth": 970, "death": "-"},
    "Biscione": {"birth": "-", "death": "-"},
    "Indra": {"birth": "-", "death": "-"},
    "Miyu Edelfelt": {"birth": 1999, "death": "-"},
    "Demeter": {"birth": "-", "death": "-"},
    "Martha": {"birth": 1, "death": 80},
    "Sakata Kintoki": {"birth": 956, "death": 1012},
    "Quetzalcoatl": {"birth": "-", "death": "-"},
    "Achilles": {"birth": -1250, "death": -1184},
    "Red Hare": {"birth": 160, "death": 200},
    "Carmilla": {"birth": 1560, "death": 1614},
    "Mandricardo": {"birth": 760, "death": "-"},
    "Europa": {"birth": -1250, "death": "-"},
    "Odysseus": {"birth": -1250, "death": -1170},
    "Nemo": {"birth": 1819, "death": "-"},
    "Habetrot": {"birth": "-", "death": "-"},
    "Huang Feihu": {"birth": -1100, "death": -1046},
    "Andromeda": {"birth": -1250, "death": "-"},
    "Noah": {"birth": -2900, "death": -1950},
    "Hippolyta": {"birth": -1250, "death": "-"},
    "Medea": {"birth": -1250, "death": "-"},
    "Mephistopheles": {"birth": 1480, "death": "-"},
    "Nursery Rhyme": {"birth": 1765, "death": "-"},
    "Irisveil": {"birth": 1980, "death": "-"},
    "Nitocris": {"birth": -2200, "death": -2180},
    "Merlin": {"birth": 450, "death": "-"},
    "Scheherazada": {"birth": 760, "death": "-"},
    "Circe": {"birth": -1250, "death": "-"},
    "Queen of Sheba": {"birth": -1020, "death": -955},
    "Sieg": {"birth": 2017, "death": "-"},
    "Skadi": {"birth": "-", "death": "-"},
    "Shuten Doji": {"birth": 950, "death": 995},
    "Asclepius": {"birth": -1250, "death": -1200},
    "Chen Gong": {"birth": 155},
    "Miss Crane": {"birth": "-", "death": "-"},
    "Daikokuten": {"birth": "-", "death": "-"},
    "Zhang Jue": {"birth": 140},
    "Aesc the Savior": {"birth": "-", "death": "-"},
    "Kuonji Alice": {"birth": 1960, "death": "-"},
    "Hanasaka no Okina": {"birth": "-", "death": "-"},
    "Hassan of Cursed Arm": {"birth": 1100, "death": "-"},
    "Stheno": {"birth": "-", "death": "-"},
    "Jing Ke": {"birth": -240},
    "Phantom of the Opera": {"birth": 1830, "death": 1881},
    "Henry Jekyll and Hyde": {"birth": 1850, "death": 1886},
    "Mysterious Heroine X": {"birth": "-", "death": "-"},
    "Hassan of the Hundred Faces": {"birth": 1100, "death": "-"},
    "Hassan of the Serenity": {"birth": 1100, "death": "-"},
    "Yan Qing": {"birth": 1080, "death": "-"},
    "Mochizuki Chiyome": {"birth": 1520, "death": 1580},
    "Semiramis": {"birth": -850, "death": -800},
    "Kama": {"birth": "-", "death": "-"},
    "Gray": {"birth": 2004, "death": "-"},
    "Kiichi Hogen": {"birth": 1100, "death": 1180},
    "Koyanskaya": {"birth": 2017, "death": "-"},
    "Thrud": {"birth": "-", "death": "-"},
    "Hildr": {"birth": "-", "death": "-"},
    "Ortlinde": {"birth": "-", "death": "-"},
    "Huyan Zhou": {"birth": 1080, "death": "-"},
    "Tezcatlipoca": {"birth": "-", "death": "-"},
    "Locusta": {"birth": 10},
    "Hassan of the Shining Star": {"birth": 1100, "death": "-"},
    "Hebi Nyobo": {"birth": "-", "death": "-"},
    "Heracles": {"birth": -1250, "death": "-"},
    "Asterios": {"birth": -1250, "death": -1200},
    "Eric Bloodaxe": {"birth": 885},
    "Beowulf": {"birth": 500, "death": 575},
    "Penthesilia": {"birth": -1250, "death": -1184},
    "Paul Bunyan": {"birth": 1834, "death": "-"},
    "Mori Nagayoshi": {"birth": 1558, "death": 1584},
    "Salome": {"birth": 14, "death": 62},
    "Kijyo Kiyo": {"birth": 920, "death": 969},
    "Galatea": {"birth": "-", "death": "-"},
    "Morgan": {"birth": 480, "death": "-"},
    "Duryodhana": {"birth": -3138, "death": -3067},
    "Shizuki Sojuro": {"birth": 1960, "death": "-"},
    "Louhi": {"birth": "-", "death": "-"},
    "Lilith": {"birth": "-", "death": "-"},
    "Sherlock Holmes": {"birth": 1854, "death": "-"},
    "Astraea": {"birth": "-", "death": "-"},
    "Amor": {"birth": "-", "death": "-"},
    "Popess Johanna": {"birth": 814},
    "Uesugi Kenshin": {"birth": 1530, "death": 1578},
    "Metatron": {"birth": "-", "death": "-"},
    "Angra Mainyu": {"birth": "-", "death": "-"},
    "Hessian Lobo": {"birth": 1893, "death": 1896},
    "Taira no Kagekiyo": {"birth": 1169},
    "Ushi Gozen": {"birth": 948, "death": 1021},
    "Elizabeth of the End": {"birth": 1560, "death": 1614},
    "BB": {"birth": 2030, "death": "-"},
    "Ganesha": {"birth": "-", "death": "-"},
    "Sesshoin Kiara": {"birth": 1994, "death": "-"},
    "Archetype Earth": {"birth": "-", "death": "-"},
    "Mysterious Executor CIEL": {"birth": 1980, "death": "-"},
    "Tenochtitlan": {"birth": 1325, "death": 1521},
    "Moon Rabbit": {"birth": "-", "death": "-"},
    "Kishinami Hakuno": {"birth": 2000, "death": "-"},
    "Meltryllis": {"birth": 2030, "death": "-"},
    "Sitonai": {"birth": "-", "death": "-"},
    "Kingprotea": {"birth": 2030, "death": "-"},
    "Ashiya Douman": {"birth": 920, "death": 1005},
    "Manannan mac Lir": {"birth": "-", "death": "-"},
    "Taisui Xingjun": {"birth": "-", "death": "-"},
    "Azumi no Isora": {"birth": "-", "death": "-"},
    "Kazuradrop": {"birth": 2030, "death": "-"},
    "Flora": {"birth": "-", "death": "-"},
    "Yang Gufei": {"birth": 719, "death": 756},
    "Kukulkan": {"birth": "-", "death": "-"},
    "Wandjina": {"birth": "-", "death": "-"},
    "Yaran-doo": {"birth": "-", "death": "-"},
    "Aozaki Aoko": {"birth": 1970, "death": "-"},
    "Mysterious Heroine XX": {"birth": "-", "death": "-"},
    "Oberon": {"birth": "-", "death": "-"},
    "Lady Avalon": {"birth": 450, "death": "-"},
    "Nine Tattoo Dragon Eliza": {"birth": 1080, "death": "-"},
    "Cait Cu Cerpriestess": {"birth": "-", "death": "-"},
    "Alessandro di Cagliostro": {"birth": 1743, "death": 1795},
    "Phantasmoon": {"birth": "-", "death": "-"},
    "Typhon Ephemeros": {"birth": "-", "death": "-"},
}


def with_blue(cell, value):
    cell.value = value
    base = cell.font.copy()
    cell.font = Font(
        name=base.name,
        sz=base.sz,
        b=base.b,
        i=base.i,
        vertAlign=base.vertAlign,
        underline=base.underline,
        strike=base.strike,
        color="0000FF",
        family=base.family,
        charset=base.charset,
        scheme=base.scheme,
        outline=base.outline,
        shadow=base.shadow,
        condense=base.condense,
        extend=base.extend,
    )


def main():
    wb = load_workbook(INPUT)
    ws = wb.active
    filled = []

    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        entry = KNOWN.get(str(name), {})
        for col, key in [(2, "birth"), (3, "death")]:
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                value = entry.get(key, "-")
                with_blue(cell, value)
                filled.append((row, name, key, value))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print(f"saved={OUTPUT}")
    print(f"filled_count={len(filled)}")
    for item in filled:
        print(item)


if __name__ == "__main__":
    main()
